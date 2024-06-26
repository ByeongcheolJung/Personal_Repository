import time, shutil, sys
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import re
from openpyxl import load_workbook


#오늘 날짜 설정
today_month = ""
if time.localtime().tm_mon<10:
        today_month = "0" + str(time.localtime().tm_mon)
else:
        today_month = str(time.localtime().tm_mon)

today_day = ""
if time.localtime().tm_mday<10:
        today_day = "0" + str(time.localtime().tm_mday)
else:
        today_day = str(time.localtime().tm_mday)


#파일 저장할때 시간 설정
today_hour = ""
if time.localtime().tm_hour<10:
        today_hour = "0" + str(time.localtime().tm_hour)
else:
        today_hour = str(time.localtime().tm_hour)

today_min = ""
if time.localtime().tm_min<10:
        today_min = "0" + str(time.localtime().tm_min)
else:
        today_min = str(time.localtime().tm_min)


#일일보고 파일 복사, 월,일,시,분 입력된상태로 만들어짐
file_path = 'C:\\Users\\Gurutech\\Desktop\\일일보고\\'
file_timestamp = today_month+today_day+'_'+today_hour+today_min
file_name = file_path+'일일보고 ('+file_timestamp+')'

try:
        os.mkdir(file_name)
        print(file_name+" 파일 생성")
except:
        print(file_name+" 파일 이미 존재. 덮어쓰기")

shutil.copy(file_path+'양식파일(복사해서 쓰세요)\\통계생성엑셀_원본v1.8.xlsx' ,file_name+'\\통계생성엑셀_원본v1.8.xlsx')

workbook = load_workbook('양식.xlsx')
worksheet = workbook.active

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
#로그인 페이지 이동
driver.get('https://gooddata.go.kr/dqe/account/login')
driver.implicitly_wait(30)
print("PMO자료 다운로드")
driver.find_element(By.ID, 'username').send_keys("pmo")
driver.find_element(By.ID, 'password').send_keys("pmo$567*")
driver.find_element(By.CLASS_NAME, 'login-form-btn').click()
driver.implicitly_wait(10)

driver.get('https://gooddata.go.kr/dqe/pmo/databases/databasesXlsDownload')
driver.implicitly_wait(20)
driver.find_element(By.XPATH, '//*[@id="navbarDropdown"]/div/span[2]').click()
driver.find_element(By.XPATH, '/html/body/div[3]/div[1]/div[3]/div/div/form/input[2]').click()
#PMO다운 파일 복사
print("PMO파일 복사 및 원본 삭제")
shutil.copy('C:\\Users\\Gurutech\\Downloads\\2022'+today_month+today_day+'_보유DB현황.xlsx' ,file_name+'\\2022'+today_month+today_day+'_보유DB현황.xlsx')
os.remove('C:\\Users\\Gurutech\\Downloads\\2022'+today_month+today_day+'_보유DB현황.xlsx')


#아이디 비밀번호 입력후 로그인 화면 이동
print("웹페이지 크롤링")
driver.find_element(By.ID, 'username').send_keys("sysMaster")
driver.find_element(By.ID, 'password').send_keys("sys$567*")
driver.find_element(By.CLASS_NAME, 'login-form-btn').click()
driver.implicitly_wait(10)

#DB확정상태 검색!!!
#driver.get('https://gooddata.go.kr/dqe/admin/databases?page=0&searchYear=2022&dbSelectionStatus=SELECTED&institutionCode=')

#대상여부 O만 뽑을떄!!!
#https://gooddata.go.kr/dqe/admin/databases?searchYear=2022&isTarget=true&isSelected=&dbSelectionStatus=&institutionType=&institutionCode=&insttName=

#전체 기관 뽑을떄!!!
#driver.get('https://gooddata.go.kr/dqe/admin/databases?page=0&searchYear=2022&')

#평가대상 기관만 뽑을떄!!!
driver.get('https://gooddata.go.kr/dqe/admin/databases?searchYear=2022&isTarget=true&isSelected=&dbSelectionStatus=&institutionType=&institutionCode=&insttName=')

#결과값 갯수
search_result = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/span').text
#print('결과값 : '+search_result)
print('결과값 : '+ re.sub(r'[^0-9]','',search_result)) #숫자만 추출
search_res = int(re.sub(r'[^0-9]','',search_result))

#데이터 임시 저장할곳
data_array = [[0]*19 for i in range(search_res)]

#페이지 갯수
pageNum = int(search_res/10)
stoppoint = True

#데이터 저장
for page in range(0,pageNum+1):
        print(str(page+1) + "P / " + str(pageNum+1) + "P")
        #확정상태 뽑을떄!!!
        #driver.get('https://gooddata.go.kr/dqe/admin/databases?page='+str(page)+'&searchYear=2022&dbSelectionStatus=SELECTED&institutionCode=')

        #대상여부가 O인 기관만 뽑을떄!!!
        #driver.get('https://gooddata.go.kr/dqe/admin/databases?page='+str(page)+'&searchYear=2022&isTarget=true&institutionCode=')

        #모든기괸 다 뽑을떄!!!
        #driver.get('https://gooddata.go.kr/dqe/admin/databases?page='+str(page)+'&searchYear=2022&')

        #평가대상인 기관만 뽑을떄!!!
        driver.get('https://gooddata.go.kr/dqe/admin/databases?page='+str(page)+'&searchYear=2022&isTarget=true&institutionCode=')

        driver.implicitly_wait(10)
        for i in range(0,10):

                lineNum = str(i+1)
                try:
                        data_array[i + page * 10][0] = i + page*10+1
                except:
                        print("크롤링 끝!!!")
                        stoppoint = False
                        break

                data_array[i+page*10][1] = driver.find_element(By.XPATH, '// *[ @ id = "app"] / div[4] / div / div[1] / div[7] / table / tbody / tr['+lineNum+'] / td[3]').text
                data_array[i+page*10][2] = driver.find_element(By.XPATH, '// *[ @ id = "app"] / div[4] / div / div[1] / div[7] / table / tbody / tr['+lineNum+'] / td[4]').text
                data_array[i+page*10][3] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[5]/a').text
                data_array[i+page*10][4] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[6]').text
                data_array[i+page*10][5] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[7]').text
                data_array[i+page*10][6] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[8]').text
                data_array[i+page*10][7] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[9]').text
                data_array[i+page*10][8] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[10]').text
                data_array[i+page*10][9] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[11]').text
                data_array[i+page*10][10] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[12]').text
                data_array[i+page*10][11] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[13]').text
                data_array[i+page*10][12] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[14]').text
                data_array[i+page*10][13] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[15]').text
                data_array[i+page*10][14] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[16]').text
                data_array[i+page*10][15] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[17]').text
                data_array[i+page*10][16] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[18]').text
                data_array[i+page*10][17] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[19]').text
                data_array[i+page*10][18] = driver.find_element(By.XPATH, '//*[@id="app"]/div[4]/div/div[1]/div[7]/table/tbody/tr['+lineNum+']/td[20]').text

        if stoppoint:
                continue
        else:
                break

#데이터 넣기
for x in range(0,search_res):
        #결과 표시
        #print(data_array[x])
        for y in range(0, 19):
                worksheet.cell(row = 3+x,column=y+1, value=data_array[x][y])

#Excel로 데이터 저장
excelname = 'CRAWLING-RESULT_'+file_timestamp+'.xlsx'
print("[저장위치] >>> "+file_name+'\\CRAWLING-RESULT_'+file_timestamp+'.xlsx')
workbook.save(file_name+'\\CRAWLING-RESULT_'+file_timestamp+'.xlsx')
workbook.close()

#경로에 있는 파일 탐색기 열기
os.startfile(file_name)
