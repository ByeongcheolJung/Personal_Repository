import os, openpyxl

#파일 위치 정하고 리스트에 저장
from openpyxl.styles import PatternFill, Font

file_list = os.listdir('C:\\Users\\Gurutech\\Desktop\\품질시스템 현행화 (2)')


wb = openpyxl.Workbook()

sheet = wb.active
sheet.column_dimensions["A"].width = 50
sheet.title = "삭제공문처리"
sheet.cell(1,1).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
sheet.cell(1,1).font = Font(color="FFFFFF")
tmp = 0

for n in range(0,len(file_list)):
    folder_list = os.listdir('C:\\Users\\Gurutech\\Desktop\\품질시스템 현행화 (2)\\'+file_list[n])
    for m in range(0, len(folder_list)):
        sheet.cell(row=tmp+2, column=1).value = folder_list[m]
        tmp = tmp + 1

sheet.cell(1,1).value = "파일명 >>> "+str(tmp)+"개의 파일"
wb.save("File_list.xlsx")
print("종료")
#os.startfile(os.path.relpath('C:\\Users\\Gurutech\\PycharmProjects\\Read_filelist'))