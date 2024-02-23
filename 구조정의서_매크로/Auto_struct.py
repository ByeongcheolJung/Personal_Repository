from openpyxl import load_workbook
import time
import os
import time

#시간 설정
today_month = ""
if time.localtime().tm_mon<10:
    today_month = "0" + str(time.localtime().tm_mon)
else:
    today_month = str(time.localtime().tm_mon)

today_day = ""
if time.localtime().tm_mday<10:
    today_day = "0" + str(time.localtime().tm_mday)
else:
    today_day = str(time.localtime().tm_mday)

today_hour = ""
if time.localtime().tm_hour<10:
    today_hour = "0" + str(time.localtime().tm_hour)
else:
    today_hour = str(time.localtime().tm_hour)

today_min = ""
if time.localtime().tm_min<10:
    today_min = "0" + str(time.localtime().tm_min)
else:
    today_min = str(time.localtime().tm_min)
file_timestamp = today_month+today_day+today_hour+today_min

#file_path = 'C:\\Users\\user\\Desktop\구조정의서_매크로\\'

print("__file__ = "+(__file__))
path = (__file__)[:(__file__).find('Auto_struct.py')]
print(path)
file_path = path

org_file = '★구조정의서 양식_매크로★.xlsx'
org_filename = file_path+org_file
copy_filename = file_path

#**********데이터베이스 정의서 복사
#DB정보 불러오기
tempfilename = 'WDQ진단대상DB관리.xlsx'
wb = load_workbook(file_path+tempfilename,data_only=True)

ws = wb.active
row_max = ws.max_row

#데이터 복사후 임시 저장
dataarray = [[0 for col in range(3)] for row in range(row_max)]
for i in range(0,row_max-1):
    dataarray[i][0] = ws.cell(row=2+i, column=10).value
    dataarray[i][1] = ws.cell(row=2+i, column=3).value
    if str(ws.cell(row=2+i, column=5).value) == "None":
        dataarray[i][2] = str(ws.cell(row=2+i, column=4).value)
    else:
        dataarray[i][2] = str(ws.cell(row=2+i, column=4).value)+str(ws.cell(row=2+i, column=5).value)
         
#저장할 파일 생성
wb.close()
gov_name = dataarray[0][0]
copy_filename += gov_name
copy_filename += "_"+file_timestamp
try:
        os.mkdir(copy_filename)
        print(copy_filename+" 파일 생성")
except:
        print(copy_filename+" 파일 이미 존재. 덮어쓰기")

wb = load_workbook(org_filename, data_only=True)
ws = wb.active
ws = wb['데이터베이스정의서']

#데이터베이스 정의서 붙여넣기
for x in range(1,len(dataarray)):
    for y in range(1,len(dataarray[0])+1):
        ws.cell(row=x+1, column=y).value = dataarray[x-1][y-1]

wb.save(copy_filename+"\\"+gov_name+'_구조정의서.xlsx')
wb.close()

#**********테이블 정의서 복사
tempfilename = '도메인규칙.xlsx'
wb = load_workbook(file_path+tempfilename,data_only=True)

ws = wb.active
row_max = ws.max_row-2
#데이터 복사후 임시 저장
dataarray = [[0 for col in range(2)] for row in range(row_max+1)]
for i in range(0,row_max):
    dataarray[i][0] = ws.cell(row=3+i, column=3).value
    dataarray[i][1] = str(ws.cell(row=3+i, column=4).value)
wb.close()

#테이블 정의서 붙여넣기
wb = load_workbook(copy_filename+"\\"+gov_name+'_구조정의서.xlsx', data_only=True)
ws = wb.active
ws = wb['테이블정의서']

k=0
cnt=0

for x in range(1,len(dataarray)+1):
    for y in range(1,3):
        ws.cell(row=k+2, column=y).value = dataarray[x-1][y-1]

    for z in range(1,x):
        if (dataarray[x-1][0] == dataarray[z-1][0]):
            #print("x="+str(x-1)+"    "+"x="+str(z-1)+"    "+"x="+str(dataarray[x-1][0])+"    "+"x="+str(dataarray[z-1][0]))
            k-=1
            break
    k+=1
#print("******x="+str(x-1)+"    "+"x="+str(z-1)+"    "+"x="+str(ws.cell(row=k+1, column=y).value)+"    "+"x="+str(dataarray[z-1][0]))

if ws.cell(row=k+1, column=1).value == 0:
    ws.cell(row=k+1, column=1).value = ""
    ws.cell(row=k+1, column=2).value = ""

wb.save(copy_filename+"\\"+gov_name+'_구조정의서.xlsx')
wb.close()


#**********컬럼정의서 복사
tempfilename = '도메인규칙.xlsx'
wb = load_workbook(file_path+tempfilename,data_only=True)
ws = wb.active
row_max = ws.max_row-2
notnull_index = row_max
#데이터 복사후 임시 저장
dataarray = [[0 for col in range(5)] for row in range(row_max+1)]
#NOT NULL여부 확인용 임시 배열
notnullarray = [0 for col in range(row_max+1)]

for i in range(0,row_max):
    for x in range(0,5):
        dataarray[i][x] = ws.cell(row=3+i, column=x+3).value
    #print(dataarray[i])
wb.close()

#테이블 정의서 붙여넣기
wb = load_workbook(copy_filename+"\\"+gov_name+'_구조정의서.xlsx', data_only=True)
ws = wb.active
ws = wb['컬럼정의서']

for x in range(0,len(dataarray)-1):
    for y in range(0,4):
        ws.cell(row=x+2, column=y+1).value = dataarray[x][y]

    #print(str(dataarray[x][4]))
    #데이터 타입 표시
    if '(' not in str(dataarray[x][4]):
        ws.cell(row=x+2, column=5).value = dataarray[x][4]
    else:
        #데이터 길이 표시
        split_data = dataarray[x][4]
        ws.cell(row=x+2, column=5).value = split_data[:split_data.find('(')]
        split_data_length = split_data[split_data.find('(')+1:len(split_data)-1]

        #소수점 길이 표시
        if ',' not in split_data_length:
            ws.cell(row=x+2, column=6).value = split_data_length
        else:
            ws.cell(row=x+2, column=6).value = split_data_length[:split_data_length.find(',')]
            ws.cell(row=x+2, column=7).value = split_data_length[split_data_length.find(',')+1:]

    #NOT NULL 배열에 테이블명 + 컬럼명으로 키값 생성
    notnullarray[x] = dataarray[x][0] + dataarray[x][2]
wb.save(copy_filename+"\\"+gov_name+'_구조정의서.xlsx')    
wb.close()


#NOT NULL 여부 확인
tempfilename = 'WDQ진단대상컬럼제외관리.xlsx'
wb = load_workbook(file_path+tempfilename,data_only=True)
ws = wb.active
row_max = ws.max_row-1
dataarray = [[0 for col in range(2)] for row in range(row_max+1)]

for i in range(0,row_max):
    dataarray[i][0] = str(ws.cell(row=2+i, column=5).value) + str(ws.cell(row=2+i, column=6).value)
    if ws.cell(row=2+i, column=8).value == 'Y':
        dataarray[i][1] = 'N'
    else:
        dataarray[i][1] = 'Y'
wb.close()

wb = load_workbook(copy_filename+"\\"+gov_name+'_구조정의서.xlsx', data_only=True)
ws = wb.active
ws = wb['컬럼정의서']

for x in range(0, notnull_index):
    for y in range(0, row_max):
        if notnullarray[x] == dataarray[y][0]:
            ws.cell(row=x+2, column=8).value = dataarray[y][1]
            break


#Excel로 데이터 저장
print("[저장위치] >>> "+copy_filename+gov_name+'_구조정의서.xlsx')
wb.save(copy_filename+"\\"+gov_name+'_구조정의서.xlsx')
wb.close()

#경로에 있는 파일 탐색기 열기
os.startfile(file_path)

time.sleep(100)